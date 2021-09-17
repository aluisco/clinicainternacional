from django.shortcuts import render
from django.views.generic import ListView
from openpyxl import load_workbook
from .models import Paciente


class ResultadosCovidListView(ListView):
    template_name = 'resultadoscovid/index.html'
    model = Paciente


def import_xlsx(request):
    if request.method == 'POST':
        pacientes = load_workbook(request.FILES['xlsfile'])
        pacientes = pacientes.active
        for paciente in range(2, pacientes.max_row):
            if pacientes.cell(column=3, row=paciente).value is not None:
                paciente_new = Paciente(
                    nombre=pacientes.cell(column=3, row=paciente).value,
                    edad=pacientes.cell(column=4, row=paciente).value,
                    sexo=pacientes.cell(column=5, row=paciente).value,
                    ci=pacientes.cell(column=6, row=paciente).value,
                    direccion=pacientes.cell(column=7, row=paciente).value,
                    municipio=pacientes.cell(column=8, row=paciente).value,
                    provincia=pacientes.cell(column=9, row=paciente).value,
                    condicion=pacientes.cell(column=10, row=paciente).value,
                    procedencia=pacientes.cell(column=11, row=paciente).value,
                    inicio_sintoma=pacientes.cell(column=12, row=paciente).value,
                    toma_muestra=pacientes.cell(column=13, row=paciente).value,
                    tipo_muestra=pacientes.cell(column=14, row=paciente).value,
                    area=pacientes.cell(column=15, row=paciente).value,
                    resultado=pacientes.cell(column=16, row=paciente).value,
                )
                paciente_new.save()
    return render(request, 'resultadoscovid/import.html')


def estadistica(request):
    context = {
    }
    return render(request, 'resultadoscovid/estadistica.html', context=context)
